from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _safe_get(data: dict | None, *keys, default=None):
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current is None:
            return default
    return current


COLUMNS = [
    "date",
    "steps",
    "step_goal",
    "distance_m",
    "calories_total",
    "calories_active",
    "calories_bmr",
    "floors_up",
    "floors_down",
    "active_seconds",
    "sedentary_seconds",
    "moderate_intensity_min",
    "vigorous_intensity_min",
    "resting_hr",
    "min_hr",
    "max_hr",
    "sleep_seconds",
    "deep_sleep_seconds",
    "light_sleep_seconds",
    "rem_sleep_seconds",
    "awake_seconds",
    "sleep_score",
    "avg_stress",
    "max_stress",
    "rest_stress_seconds",
    "low_stress_seconds",
    "medium_stress_seconds",
    "high_stress_seconds",
    "body_battery_high",
    "body_battery_low",
    "body_battery_charged",
    "body_battery_drained",
    "spo2_avg",
    "spo2_lowest",
    "respiration_waking",
    "respiration_sleep",
    "respiration_highest",
    "respiration_lowest",
    "weight_kg",
    "bmi",
    "body_fat_pct",
    "muscle_mass_kg",
    "hydration_ml",
    "hydration_goal_ml",
    "hrv_last_night",
    "hrv_weekly_avg",
    "hrv_status",
    "training_readiness",
    "training_readiness_level",
    "vo2max",
    "training_load_7d",
    "training_status",
    "fitness_age",
    "chronological_age",
]

ACTIVITY_COLUMNS = [
    "activity_count",
    "activity_sports",
    "activity_distance_m",
    "activity_duration_sec",
    "activity_calories",
    "activity_training_load",
    "activity_aerobic_te",
    "activity_anaerobic_te",
    "primary_sport",
    "primary_name",
    "primary_start",
    "primary_distance_m",
    "primary_duration_sec",
    "primary_moving_sec",
    "primary_avg_hr",
    "primary_max_hr",
    "primary_calories",
    "primary_avg_speed_mps",
    "primary_max_speed_mps",
    "primary_elevation_gain_m",
    "primary_elevation_loss_m",
    "primary_avg_cadence",
    "primary_avg_power",
    "primary_norm_power",
    "primary_training_load",
    "primary_te_label",
    "primary_hr_zone1_sec",
    "primary_hr_zone2_sec",
    "primary_hr_zone3_sec",
    "primary_hr_zone4_sec",
    "primary_hr_zone5_sec",
    "primary_lap_count",
    "primary_location",
]

COLUMNS = COLUMNS + ACTIVITY_COLUMNS


def _activity_values(data: dict) -> list:
    """Aggregate activity metrics for the day plus detail for the primary activity.

    Primary = longest activity by duration. Prefers the rich `activities`
    (get_activities_by_date) source; falls back to the lean `activities_for_date`.
    """
    acts = data.get("activities")
    if not isinstance(acts, list) or not acts:
        acts = data.get("activities_for_date")
    acts = [a for a in acts if isinstance(a, dict)] if isinstance(acts, list) else []

    if not acts:
        return [None] * len(ACTIVITY_COLUMNS)

    sports: list[str] = []
    for a in acts:
        sp = _safe_get(a, "activityType", "typeKey")
        if sp and sp not in sports:
            sports.append(sp)

    def _sum(key):
        vals = [a.get(key) for a in acts if a.get(key) is not None]
        return round(sum(vals), 2) if vals else None

    def _max(key):
        vals = [a.get(key) for a in acts if a.get(key) is not None]
        return max(vals) if vals else None

    p = max(acts, key=lambda a: a.get("duration") or 0)

    return [
        len(acts),
        ", ".join(sports) if sports else None,
        _sum("distance"),
        _sum("duration"),
        _sum("calories"),
        _sum("activityTrainingLoad"),
        _max("aerobicTrainingEffect"),
        _max("anaerobicTrainingEffect"),
        _safe_get(p, "activityType", "typeKey"),
        p.get("activityName"),
        p.get("startTimeLocal"),
        p.get("distance"),
        p.get("duration"),
        p.get("movingDuration"),
        p.get("averageHR"),
        p.get("maxHR"),
        p.get("calories"),
        p.get("averageSpeed"),
        p.get("maxSpeed"),
        p.get("elevationGain"),
        p.get("elevationLoss"),
        p.get("averageRunningCadenceInStepsPerMinute"),
        p.get("avgPower"),
        p.get("normPower"),
        p.get("activityTrainingLoad"),
        p.get("trainingEffectLabel"),
        p.get("hrTimeInZone_1"),
        p.get("hrTimeInZone_2"),
        p.get("hrTimeInZone_3"),
        p.get("hrTimeInZone_4"),
        p.get("hrTimeInZone_5"),
        p.get("lapCount"),
        p.get("locationName"),
    ]


def _extract_row(data: dict, target_date: date) -> list:
    d = target_date.isoformat()
    summary = data.get("user_summary") or data.get("stats") or {}
    hr = data.get("heart_rates") or {}
    sleep_dto = _safe_get(data, "sleep", "dailySleepDTO") or {}
    stress = data.get("stress_all_day") or data.get("stress_detailed") or {}
    bb_list = data.get("body_battery")
    body = data.get("body_composition") or {}
    spo2 = data.get("spo2") or {}
    resp = data.get("respiration") or {}
    hydration = data.get("hydration") or {}
    hrv_summary = _safe_get(data, "hrv", "hrvSummary") or data.get("hrv") or {}
    tr = data.get("training_readiness") or {}
    if isinstance(tr, list):
        tr = tr[0] if tr else {}
    ts = data.get("training_status") or {}
    fa = data.get("fitness_age") or {}

    weight = body.get("weight")
    if weight and weight > 500:
        weight = weight / 1000
    muscle = body.get("muscleMass")
    if muscle and muscle > 500:
        muscle = muscle / 1000

    bb = None
    if bb_list and isinstance(bb_list, list):
        for entry in bb_list:
            if entry and (entry.get("calendarDate") == d or entry.get("date") == d):
                bb = entry
                break
        if not bb:
            bb = bb_list[-1] if bb_list else None
    bb = bb or {}

    bb_high = bb.get("bodyBatteryHighValue") or bb.get("highest")
    bb_low = bb.get("bodyBatteryLowValue") or bb.get("lowest")
    if not bb_high and bb.get("bodyBatteryValuesArray"):
        vals = [v[1] for v in bb["bodyBatteryValuesArray"] if v and len(v) > 1 and v[1] is not None]
        if vals:
            bb_high = max(vals)
            bb_low = min(vals)

    return [
        d,
        summary.get("totalSteps"),
        summary.get("dailyStepGoal"),
        summary.get("totalDistanceMeters"),
        summary.get("totalKilocalories"),
        summary.get("activeKilocalories"),
        summary.get("bmrKilocalories"),
        summary.get("floorsAscended"),
        summary.get("floorsDescended"),
        summary.get("activeSeconds"),
        summary.get("sedentarySeconds"),
        summary.get("moderateIntensityMinutes"),
        summary.get("vigorousIntensityMinutes"),
        hr.get("restingHeartRate"),
        hr.get("minHeartRate"),
        hr.get("maxHeartRate"),
        sleep_dto.get("sleepTimeSeconds"),
        sleep_dto.get("deepSleepSeconds"),
        sleep_dto.get("lightSleepSeconds"),
        sleep_dto.get("remSleepSeconds"),
        sleep_dto.get("awakeSleepSeconds"),
        _safe_get(sleep_dto, "sleepScores", "overall", "value") or sleep_dto.get("sleepScore"),
        stress.get("avgStressLevel") or stress.get("overallStressLevel"),
        stress.get("maxStressLevel"),
        stress.get("restStressDuration"),
        stress.get("lowStressDuration"),
        stress.get("mediumStressDuration"),
        stress.get("highStressDuration"),
        bb_high,
        bb_low,
        bb.get("charged"),
        bb.get("drained"),
        _safe_get(spo2, "averageSpO2") or _safe_get(spo2, "dailySpO2Values", "averageSpO2"),
        _safe_get(spo2, "lowestSpO2") or _safe_get(spo2, "dailySpO2Values", "lowestSpO2"),
        resp.get("avgWakingRespirationValue"),
        resp.get("avgSleepRespirationValue"),
        resp.get("highestRespirationValue"),
        resp.get("lowestRespirationValue"),
        weight,
        body.get("bmi"),
        body.get("bodyFat"),
        muscle,
        hydration.get("valueInML") or hydration.get("intakeinML"),
        hydration.get("goalInML"),
        hrv_summary.get("lastNight") or hrv_summary.get("lastNightAvg"),
        hrv_summary.get("weeklyAvg") or hrv_summary.get("weeklyAverage"),
        hrv_summary.get("status") or hrv_summary.get("currentStatus"),
        tr.get("score") or _safe_get(tr, "trainigReadinessDTO", "score"),
        tr.get("level") or _safe_get(tr, "trainigReadinessDTO", "level"),
        ts.get("vo2MaxValue") or _safe_get(ts, "mostRecentVO2Max", "generic", "vo2MaxValue"),
        ts.get("trainingLoad7Day"),
        ts.get("trainingStatusPhrase") or ts.get("currentTrainingStatus"),
        fa.get("fitnessAge"),
        fa.get("chronologicalAge"),
    ] + _activity_values(data)


def _hm(seconds) -> str:
    if not seconds:
        return "N/A"
    s = int(seconds)
    return f"{s // 3600}h {(s % 3600) // 60}m"


def _km(meters) -> str:
    if not meters:
        return "N/A"
    return f"{meters / 1000:.1f} km"


def format_summary(data: dict, target_date: date) -> str:
    row = _extract_row(data, target_date)
    v = dict(zip(COLUMNS, row))
    weekday = target_date.strftime("%A")

    steps = v["steps"] or 0
    goal = v["step_goal"] or 0
    pct = round(steps / goal * 100) if goal else 0

    lines = [
        f"Garmin {v['date']} ({weekday})",
        "",
        f"Steps: {steps:,} / {goal:,} ({pct}%)",
        f"Distance: {_km(v['distance_m'])}",
        f"Calories: {int(v['calories_total'] or 0)} ({int(v['calories_active'] or 0)} active)",
        f"Floors: {int(v['floors_up'] or 0)} up / {int(v['floors_down'] or 0)} down",
        "",
        f"Resting HR: {v['resting_hr'] or 'N/A'} bpm",
        f"HR range: {v['min_hr'] or '?'} - {v['max_hr'] or '?'} bpm",
        f"HRV: {v['hrv_last_night'] or 'N/A'} ms (7d avg: {v['hrv_weekly_avg'] or 'N/A'})",
        "",
        f"Sleep: {_hm(v['sleep_seconds'])} (score: {v['sleep_score'] or 'N/A'})",
        f"  Deep {_hm(v['deep_sleep_seconds'])} / Light {_hm(v['light_sleep_seconds'])} / REM {_hm(v['rem_sleep_seconds'])}",
        "",
        f"Stress: avg {v['avg_stress'] or 'N/A'} / max {v['max_stress'] or 'N/A'}",
        f"Body Battery: {v['body_battery_low'] or '?'} - {v['body_battery_high'] or '?'} (+{v['body_battery_charged'] or '?'} / -{v['body_battery_drained'] or '?'})",
        f"SpO2: {v['spo2_avg'] or 'N/A'}% (low: {v['spo2_lowest'] or 'N/A'}%)",
    ]

    if v["training_readiness"]:
        lines.append(f"Training readiness: {v['training_readiness']} ({v['training_readiness_level'] or ''})")
    if v["vo2max"]:
        lines.append(f"VO2 Max: {v['vo2max']}")
    if v["weight_kg"]:
        lines.append(f"Weight: {v['weight_kg']:.1f} kg")

    if v.get("activity_count"):
        lines.append("")
        lines.append(f"Activities: {v['activity_count']} ({v.get('activity_sports') or ''})")
        load = v.get("primary_training_load")
        lines.append(
            f"  {v.get('primary_name') or v.get('primary_sport') or 'activity'}: "
            f"{_hm(v.get('primary_duration_sec'))}, {_km(v.get('primary_distance_m'))}, "
            f"{int(v.get('primary_calories') or 0)} kcal"
            + (f", load {round(load)}" if load else "")
        )

    return "\n".join(lines)


def _parse_text_doc(text: str) -> dict[str, dict[str, str]]:
    """Parse text doc into {date_str: {prop: value}}."""
    entries: dict[str, dict[str, str]] = {}
    current_date = None
    for line in text.splitlines():
        line = line.rstrip()
        if not line:
            continue
        if not line.startswith(" ") and ":" not in line:
            current_date = line.strip()
            entries[current_date] = {}
        elif current_date and ": " in line:
            key, value = line.split(": ", 1)
            entries[current_date][key.strip()] = value.strip()
    return entries


def append_to_text_doc(data: dict, target_date: date, txt_path: Path) -> Path:
    entries: dict[str, dict[str, str]] = {}
    if txt_path.exists():
        entries = _parse_text_doc(txt_path.read_text(encoding="utf-8"))

    row = _extract_row(data, target_date)
    day_props = {}
    for col, val in zip(COLUMNS[1:], row[1:]):  # skip "date"
        if val is not None and val != "":
            day_props[col] = str(val)

    entries[target_date.isoformat()] = day_props

    lines = []
    for date_str in sorted(entries.keys()):
        lines.append(date_str)
        for prop, value in entries[date_str].items():
            lines.append(f"  {prop}: {value}")
        lines.append("")

    txt_path.write_text("\n".join(lines), encoding="utf-8")
    return txt_path


def append_to_excel(data: dict, target_date: date, output_dir: Path, filename: str | None = None) -> Path:
    xlsx_path = output_dir / (filename or f"{target_date.year}-garmin.xlsx")

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        if header != COLUMNS:
            for idx, col in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=idx, value=col)
        existing_dates = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing_dates.add(str(row[0]))
        if target_date.isoformat() in existing_dates:
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
                if str(row[0]) == target_date.isoformat():
                    ws.delete_rows(idx)
                    break
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Garmin"
        ws.append(COLUMNS)

    ws.append(_extract_row(data, target_date))
    wb.save(xlsx_path)
    return xlsx_path
